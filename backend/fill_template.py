
import json
from pathlib import Path
from openpyxl import load_workbook

def fill_template(blank_template_path: str, mapping_csv_path: str, answers_json_path: str, output_path: str):
    """
    Fill an Excel template based on a mapping CSV and an answers JSON.
    - blank_template_path: path to the blank .xlsx template
    - mapping_csv_path: CSV with columns: sheet, cell, text, is_question, json_key
    - answers_json_path: JSON object mapping json_key -> value
    - output_path: path to write the filled workbook
    """
    import pandas as pd

    # Load mapping and answers
    df = pd.read_csv(mapping_csv_path)
    with open(answers_json_path, "r") as f:
        answers = json.load(f)

    wb = load_workbook(blank_template_path)
    # For each row where is_question == 'yes' (case-insensitive), write the answer into the cell.
    for _, row in df.iterrows():
        if str(row.get("is_question", "")).strip().lower() in ("yes", "y", "true", "1"):
            sheet = row["sheet"]
            cell = row["cell"]
            key = str(row.get("json_key", "")).strip()
            if not key:
                continue
            value = answers.get(key, None)
            # If no answer provided, skip writing (or write empty)
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                ws[cell].value = value

    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--template", required=True)
    p.add_argument("--mapping", required=True)
    p.add_argument("--answers", required=True)
    p.add_argument("--out", required=True)
    args = p.parse_args()
    out = fill_template(args.template, args.mapping, args.answers, args.out)
    print("Wrote:", out)
