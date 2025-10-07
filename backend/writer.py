from __future__ import annotations
from pathlib import Path
from typing import Any, Dict
from datetime import datetime
from openpyxl import load_workbook
import logging

def _coerce_confidence(val: Any) -> int:
  """Coerce a confidence value to an integer 1..10. Defaults to 3 if missing/invalid."""
  try:
    if val is None:
      return 3
    if isinstance(val, (int, float)):
      n = int(round(float(val)))
    else:
      s = str(val).strip()
      if not s:
        return 3
      # extract leading number if present
      num = "".join(ch for ch in s if (ch.isdigit() or ch == "."))
      n = int(round(float(num))) if num else 3
    if n < 1:
      return 1
    if n > 10:
      return 10
    return n
  except Exception:
    return 3

def _extract_structured(answer_value: Any) -> tuple[Any, int, str]:
  """Accepts either a scalar answer or a dict with keys like
  {answer|value|text, confidence, source|page|source_page|source_pages}.
  Returns (value, confidence:int 1..10, source:str)."""
  # Defaults
  conf = 3
  source = "Unknown"

  if isinstance(answer_value, dict):
    # value
    val = (answer_value.get("answer") or
           answer_value.get("value") or
           answer_value.get("text") or
           answer_value.get("result") or
           None)
    if val is None:
      val = "Unknown"
    # confidence
    conf = _coerce_confidence(answer_value.get("confidence"))
    # source / page(s)
    src = (answer_value.get("source") or
           answer_value.get("page") or
           answer_value.get("source_page") or
           answer_value.get("source_pages") or
           None)
    if src is None or src == "":
      source = "Unknown"
    elif isinstance(src, (list, tuple)):
      # Join multiple pages
      source = ", ".join(f"Page {p}" if isinstance(p, (int, float)) else str(p) for p in src)
    elif isinstance(src, (int, float)):
      source = f"Page {int(src)}"
    else:
      source = str(src)
    return val, conf, source

  # Non-dict: just a plain value
  val = answer_value if answer_value is not None else "Unknown"
  return val, conf, source

def _write_conf_source(ws, row_idx: int, conf: int, source: str):
  """Write Confidence (col C) and Source (col D)."""
  # Ensure confidence between 1 and 10, default 3
  if not isinstance(conf, int):
    conf = 3
  if conf < 1:
    conf = 1
  if conf > 10:
    conf = 10
  # Ensure source is a non-empty string
  if source is None or str(source).strip() == "":
    source = "Unknown"
  ws.cell(row=row_idx, column=3, value=conf)
  ws.cell(row=row_idx, column=4, value=source)

from .mapping import Mapping

def _try_parse_number(value: Any) -> Any:
  if value is None:
    return None
  if isinstance(value, (int, float)):
    return value
  s = str(value).strip()
  if not s:
    return None
  # Remove currency symbols and commas
  s = s.replace("$", "").replace(",", "").replace("%", "")
  try:
    if "." in s:
      return float(s)
    return int(s)
  except Exception:
    return value  # fall back to original

def _try_parse_date(value: Any):
  """Return a python date for common formats so Excel can format it nicely, else return original."""
  if value is None:
    return None
  if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
    return value
  s = str(value).strip()
  if not s:
    return None
  for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%m-%d-%Y"):
    try:
      return datetime.strptime(s, fmt).date()
    except Exception:
      continue
  # If parsing fails, just return the original string
  return s

def _normalize_yesno(value: Any) -> str | None:
  if value is None:
    return None
  s = str(value).strip().lower()
  if s in ("yes", "y", "true", "1"):
    return "Yes"
  if s in ("no", "n", "false", "0"):
    return "No"
  # Leave as-is if it's already something like "Yes"/"No" with different casing
  if s:
    return "Yes" if s == "yes" else ("No" if s == "no" else str(value))
  return None

def _coerce_for_cell(answer_value: Any, answer_type: str):
  at = (answer_type or "text").lower().strip()
  # Filter out "null" strings
  if isinstance(answer_value, str) and answer_value.strip().lower() == "null":
    return None
  if at == "date":
    return _try_parse_date(answer_value)
  if at in ("number", "currency"):
    return _try_parse_number(answer_value)
  if at == "yesno":
    return _normalize_yesno(answer_value)
  if at == "list":
    if isinstance(answer_value, list):
      return ", ".join([str(x) for x in answer_value])
    return str(answer_value) if answer_value is not None else None
  # text, email, phone, default
  return answer_value if answer_value is not None else None

def _targets(row):
  """
  Given a mapping row, determine the worksheet, row index, and target columns for question and answer.
  Returns (ws, row_idx, col_q, col_a, col_conf, col_src)
  """
  # Pick sheet (fallback to first sheet if mapping name not found)
  wb = row._workbook  # must be set by fill_template
  sheet_name = row.sheet if row.sheet in wb.sheetnames else wb.sheetnames[0]
  ws = wb[sheet_name]
  # Parse row.cell to get row index
  from openpyxl.utils.cell import coordinate_from_string
  cell = row.cell
  if not cell:
    return ws, None, None, None, None, None
  col, idx = coordinate_from_string(cell)
  row_idx = idx
  # Column A: question, B: answer, C: Confidence, D: Source
  return ws, row_idx, 1, 2, 3, 4

def _write_text(ws, row_idx, value, answer_type):
  cell = ws.cell(row=row_idx, column=2)
  val = _coerce_for_cell(value, answer_type)
  cell.value = val
  if answer_type and answer_type.lower().strip() == "date":
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
      cell.number_format = "mm/dd/yyyy"

def fill_template(mapping: Mapping, answers: Dict[str, Any], out_path: Path, excel_template: Path | None = None) -> Path:
  """
  Loads the Excel template (from `excel_template` if provided, otherwise from config.EXCEL_TEMPLATE),
  preserves all question text in column A,
  and writes answers into column B for text/date/number/percent, or C/D for yes/no.
  """
  from .config import EXCEL_TEMPLATE
  tpl = excel_template or EXCEL_TEMPLATE
  wb = load_workbook(tpl, data_only=False, keep_vba=False)

  # Attach workbook to each mapping row for _targets
  for row in mapping.question_rows:
    row._workbook = wb

  for row in mapping.question_rows:
    key = row.json_key
    if not key:
      logging.warning(f"Skipping row with no json_key at {row.cell}")
      continue
    ws, row_idx, col_q, col_a, col_conf, col_src = _targets(row)
    if ws is None or row_idx is None:
      continue
    # Preserve original question text in column A
    orig_q = ws.cell(row=row_idx, column=col_q).value
    ws.cell(row=row_idx, column=col_q, value=orig_q)

    raw_value = answers.get(key, None)
    value, conf, src = _extract_structured(raw_value)

    # Ensure no literal "null" written for answer
    answer_type = (row.answer_type or "text").lower().strip()

    try:
      _write_text(ws, row_idx, value, answer_type)
      # Confidence & Source for every row, always write even if empty
      _write_conf_source(ws, row_idx, conf, src)
    except Exception as e:
      logging.warning(f"Failed to write answer for key {key} at {row.cell}: {e}")
      ws.cell(row=row_idx, column=col_a, value=f"Error: {e}")
      ws.cell(row=row_idx, column=col_conf, value="Error")
      ws.cell(row=row_idx, column=col_src, value="Error")
      continue

  # Ensure parent directory exists and save
  out_path = Path(out_path)
  out_path.parent.mkdir(parents=True, exist_ok=True)
  wb.save(out_path)
  return out_path